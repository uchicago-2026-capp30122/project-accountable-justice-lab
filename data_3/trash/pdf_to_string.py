import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import requests
import subprocess
import tempfile
import json
import re

# === INPUTS ===
XLSX_PATH = Path("compiled_dataset_filtrado.xlsx")   # your excel file
OUT_CSV   = Path("compiled_dataset_with_pdf_words.csv")

# Column AF in Excel = 32nd column => 1-based 32, 0-based index 31
AF_COL_1BASED = 32

NEW_COL_NAME = "pdf_words"          # column with list of words
LINK_COL_NAME = "pdf_link_extracted"  # optional: keep extracted link

# --- helpers ---
def extract_hyperlinks_from_column(xlsx_path: Path, col_1based: int) -> list:
    """
    Returns a list (len = number of data rows) with the hyperlink target
    (if present) or the cell value otherwise, for a given Excel column.
    Assumes headers are in row 1 and data starts row 2.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    links = []
    # iterate data rows (row 2..max_row)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_1based)
        if cell.hyperlink and cell.hyperlink.target:
            links.append(cell.hyperlink.target)
        else:
            links.append(cell.value)
    return links

def pdftotext_words(pdf_bytes: bytes) -> list[str]:
    """
    Writes PDF bytes to a temp file, runs pdftotext, returns list of words.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        pdf_path = tmpdir / "doc.pdf"
        txt_path = tmpdir / "doc.txt"
        pdf_path.write_bytes(pdf_bytes)

        # Run pdftotext
        # -layout optional; remove it if you prefer plain extraction
        subprocess.run(
            ["pdftotext", "-layout", str(pdf_path), str(txt_path)],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        text = txt_path.read_text(errors="ignore")

    # Split into words (basic tokenization)
    # If you want ONLY letters/numbers: use regex below
    words = re.findall(r"\S+", text)
    return words

def get_pdf_bytes(link_or_path) -> bytes | None:
    """
    Accepts a URL (http/https) or local file path; returns PDF bytes or None.
    """
    if link_or_path is None:
        return None

    s = str(link_or_path).strip()
    if not s:
        return None

    # URL
    if s.lower().startswith(("http://", "https://")):
        resp = requests.get(s, timeout=60)
        resp.raise_for_status()
        return resp.content

    # Local path
    p = Path(s)
    if p.exists() and p.is_file():
        return p.read_bytes()

    return None

# === main ===
print("Reading Excel into DataFrame...")
df = pd.read_excel(XLSX_PATH, engine="openpyxl")

print("Extracting hyperlinks/values from column AF...")
af_links = extract_hyperlinks_from_column(XLSX_PATH, AF_COL_1BASED)

# Sanity check length match
if len(af_links) != len(df):
    raise RuntimeError(f"Row mismatch: df has {len(df)} rows but AF links has {len(af_links)}.")

df[LINK_COL_NAME] = af_links

print("Processing PDFs with pdftotext...")
all_words = []
for i, link in enumerate(df[LINK_COL_NAME], start=1):
    try:
        pdf_bytes = get_pdf_bytes(link)
        if pdf_bytes is None:
            all_words.append([])
            continue

        words = pdftotext_words(pdf_bytes)
        all_words.append(words)

    except subprocess.CalledProcessError:
        # pdftotext failed on this PDF
        all_words.append([])
    except Exception as e:
        # download/read error, etc.
        all_words.append([])

    if i % 50 == 0:
        print(f"  processed {i}/{len(df)}")

# Store lists safely in CSV as JSON strings
df[NEW_COL_NAME] = [json.dumps(w, ensure_ascii=False) for w in all_words]

print("Saving CSV...")
df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
print("Done:", OUT_CSV.resolve())

